"""documents.py — PDF, Excel, OCR, Word processing for Legion.

Requirements:
    pip install openpyxl pdfplumber pytesseract python-docx
    sudo apt install tesseract-ocr tesseract-ocr-eng tesseract-ocr-ind
"""

from __future__ import annotations

import asyncio
import glob
import logging
import os
import shutil
import time
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# ── PDF ──────────────────────────────────────────────────────────────────────

async def read_pdf(path: str, pages: str = "all", max_chars: int = 8000) -> str:
    """Extract text from a PDF file.

    pages: 'all', '1-5', '3', '1,3,5'
    """
    p = Path(path).expanduser()
    if not p.exists():
        return f"file not found: {path}"

    def _extract():
        import pdfplumber
        with pdfplumber.open(str(p)) as pdf:
            total_pages = len(pdf.pages)
            page_nums = _parse_page_range(pages, total_pages)

            texts = []
            for i in page_nums:
                if 0 <= i < total_pages:
                    page_text = pdf.pages[i].extract_text() or ""
                    if page_text.strip():
                        texts.append(f"--- Page {i + 1} ---\n{page_text}")

            full = "\n\n".join(texts)
            if len(full) > max_chars:
                full = full[:max_chars] + f"\n\n[...truncated, {len(full)} total chars]"
            return f"PDF: {p.name} ({total_pages} pages)\n\n{full}"

    return await asyncio.get_event_loop().run_in_executor(None, _extract)


async def pdf_extract_tables(path: str, pages: str = "all") -> str:
    """Extract tables from PDF as markdown."""
    p = Path(path).expanduser()
    if not p.exists():
        return f"file not found: {path}"

    def _extract():
        import pdfplumber
        with pdfplumber.open(str(p)) as pdf:
            total_pages = len(pdf.pages)
            page_nums = _parse_page_range(pages, total_pages)
            tables_found = []

            for i in page_nums:
                if 0 <= i < total_pages:
                    page_tables = pdf.pages[i].extract_tables()
                    for j, table in enumerate(page_tables):
                        if not table:
                            continue
                        # Convert to markdown
                        md_lines = []
                        for row_idx, row in enumerate(table):
                            cells = [str(c or "").strip() for c in row]
                            md_lines.append("| " + " | ".join(cells) + " |")
                            if row_idx == 0:
                                md_lines.append("|" + "|".join(["---"] * len(cells)) + "|")
                        tables_found.append(
                            f"Table {len(tables_found) + 1} (page {i + 1}):\n"
                            + "\n".join(md_lines)
                        )

            if not tables_found:
                return "No tables found in PDF."
            return "\n\n".join(tables_found)

    return await asyncio.get_event_loop().run_in_executor(None, _extract)


# ── Excel ────────────────────────────────────────────────────────────────────

async def read_excel(path: str, sheet: str = "", max_rows: int = 100) -> str:
    """Read Excel file, return as markdown table."""
    p = Path(path).expanduser()
    if not p.exists():
        return f"file not found: {path}"

    def _read():
        import openpyxl
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        ws = wb[sheet] if sheet and sheet in sheet_names else wb.active
        ws_name = ws.title

        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                rows.append(f"[...truncated at {max_rows} rows]")
                break
            cells = [str(c if c is not None else "") for c in row]
            rows.append("| " + " | ".join(cells) + " |")
            if i == 0:
                rows.append("|" + "|".join(["---"] * len(cells)) + "|")

        wb.close()
        header = f"Excel: {p.name} | Sheet: {ws_name} | Sheets: {', '.join(sheet_names)}\n\n"
        return header + "\n".join(rows)

    return await asyncio.get_event_loop().run_in_executor(None, _read)


async def write_excel(
    path: str,
    data: list[list[str]],
    sheet_name: str = "Sheet1",
) -> str:
    """Write data to an Excel file. data is a list of rows (each row is a list of cell values)."""
    p = Path(path).expanduser()

    def _write():
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        for row in data:
            ws.append(row)
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(p))
        return f"wrote {len(data)} rows to {p}"

    return await asyncio.get_event_loop().run_in_executor(None, _write)


async def excel_update_cell(
    path: str,
    sheet: str,
    cell: str,
    value: str,
) -> str:
    """Update a specific cell in an Excel file. cell format: 'B5', 'A1', etc."""
    p = Path(path).expanduser()
    if not p.exists():
        return f"file not found: {path}"

    def _update():
        import openpyxl
        wb = openpyxl.load_workbook(str(p))
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        ws[cell] = value
        wb.save(str(p))
        return f"updated {ws.title}!{cell} = '{value}'"

    return await asyncio.get_event_loop().run_in_executor(None, _update)


# ── OCR ──────────────────────────────────────────────────────────────────────

async def ocr_image(path: str, lang: str = "eng") -> str:
    """OCR an image file using Tesseract."""
    p = Path(path).expanduser()
    if not p.exists():
        return f"file not found: {path}"

    def _ocr():
        import pytesseract
        from PIL import Image
        img = Image.open(str(p))
        text = pytesseract.image_to_string(img, lang=lang)
        return f"OCR result ({p.name}, lang={lang}):\n\n{text.strip()}"

    return await asyncio.get_event_loop().run_in_executor(None, _ocr)


async def ocr_pdf(path: str, lang: str = "eng", pages: str = "all") -> str:
    """OCR a scanned (image-based) PDF. Converts pages to images then runs Tesseract."""
    p = Path(path).expanduser()
    if not p.exists():
        return f"file not found: {path}"

    def _ocr_pdf():
        import pdfplumber
        import pytesseract
        from PIL import Image

        texts = []
        with pdfplumber.open(str(p)) as pdf:
            total = len(pdf.pages)
            page_nums = _parse_page_range(pages, total)
            for i in page_nums:
                if 0 <= i < total:
                    page = pdf.pages[i]
                    # Try text extraction first
                    text = page.extract_text()
                    if text and text.strip():
                        texts.append(f"--- Page {i + 1} (text) ---\n{text}")
                    else:
                        # Fall back to OCR on page image
                        img = page.to_image(resolution=200)
                        pil_img = img.original
                        ocr_text = pytesseract.image_to_string(pil_img, lang=lang)
                        if ocr_text.strip():
                            texts.append(f"--- Page {i + 1} (OCR) ---\n{ocr_text}")

        if not texts:
            return "No text could be extracted from the PDF."
        return "\n\n".join(texts)

    return await asyncio.get_event_loop().run_in_executor(None, _ocr_pdf)


# ── Word documents ───────────────────────────────────────────────────────────

async def read_docx(path: str, max_chars: int = 8000) -> str:
    """Read a Word document (.docx)."""
    p = Path(path).expanduser()
    if not p.exists():
        return f"file not found: {path}"

    def _read():
        from docx import Document
        doc = Document(str(p))
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        text = "\n\n".join(paragraphs)
        if len(text) > max_chars:
            text = text[:max_chars] + f"\n\n[...truncated, {len(text)} total chars]"
        return f"Word: {p.name} ({len(paragraphs)} paragraphs)\n\n{text}"

    return await asyncio.get_event_loop().run_in_executor(None, _read)


# ── File management ──────────────────────────────────────────────────────────

async def organize_files(directory: str, strategy: str = "by_type") -> str:
    """Organize files in a directory.

    strategy: 'by_type' (group by extension), 'by_date' (by modified year-month)
    """
    d = Path(directory).expanduser()
    if not d.is_dir():
        return f"not a directory: {directory}"

    files = [f for f in d.iterdir() if f.is_file()]
    if not files:
        return "no files to organize"

    moved = []
    for f in files:
        if strategy == "by_type":
            ext = f.suffix.lower().lstrip(".") or "no_extension"
            dest_dir = d / ext
        elif strategy == "by_date":
            mtime = time.localtime(f.stat().st_mtime)
            dest_dir = d / f"{mtime.tm_year}-{mtime.tm_mon:02d}"
        else:
            return f"unknown strategy: {strategy}"

        dest_dir.mkdir(exist_ok=True)
        dest = dest_dir / f.name
        if not dest.exists():
            shutil.move(str(f), str(dest))
            moved.append(f"{f.name} → {dest_dir.name}/")

    return f"organized {len(moved)} files ({strategy}):\n" + "\n".join(moved[:30])


async def find_files(
    directory: str,
    pattern: str,
    max_depth: int = 5,
    max_results: int = 50,
) -> str:
    """Find files matching a glob pattern."""
    d = Path(directory).expanduser()
    if not d.is_dir():
        return f"not a directory: {directory}"

    matches = []
    for f in d.rglob(pattern):
        # Rough depth check
        rel = f.relative_to(d)
        if len(rel.parts) > max_depth:
            continue
        size = f.stat().st_size if f.is_file() else 0
        size_str = _format_size(size)
        matches.append(f"  {rel}  ({size_str})")
        if len(matches) >= max_results:
            matches.append(f"  [...{max_results} limit reached]")
            break

    if not matches:
        return f"no files matching '{pattern}' in {directory}"
    return f"Found {len(matches)} matches for '{pattern}':\n" + "\n".join(matches)


async def file_info(path: str) -> str:
    """Get detailed file information."""
    p = Path(path).expanduser()
    if not p.exists():
        return f"not found: {path}"

    stat = p.stat()
    info = {
        "name": p.name,
        "path": str(p),
        "type": "directory" if p.is_dir() else p.suffix or "no extension",
        "size": _format_size(stat.st_size),
        "modified": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(stat.st_mtime)),
        "created": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(stat.st_ctime)),
        "permissions": oct(stat.st_mode)[-3:],
    }
    return "\n".join(f"  {k}: {v}" for k, v in info.items())


# ── Helpers ──────────────────────────────────────────────────────────────────

def _parse_page_range(pages: str, total: int) -> list[int]:
    """Parse page range string into list of 0-indexed page numbers."""
    if pages == "all":
        return list(range(total))

    result = []
    for part in pages.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-", 1)
            start = max(1, int(start))
            end = min(total, int(end))
            result.extend(range(start - 1, end))
        else:
            n = int(part)
            if 1 <= n <= total:
                result.append(n - 1)
    return result


def _format_size(size_bytes: int) -> str:
    """Format bytes to human-readable size."""
    for unit in ["B", "KB", "MB", "GB"]:
        if size_bytes < 1024:
            return f"{size_bytes:.1f}{unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f}TB"
